from abc import ABC, abstractmethod
import os
from typing import Any

import openpyxl

class TableEngine(ABC):
    """Interface für Tabellen-Engines."""

    # Getter
    @abstractmethod
    def get_max_row(self) -> int:
        """Gibt die maximale Zeilenanzahl zurück."""
        pass
    @abstractmethod
    def get_max_col(self) -> int:
        """Gibt die maximale Spaltenanzahl zurück."""
        pass
    @abstractmethod
    def get_cell_value(self, row: int, col: int):
        """Gibt den Wert der Zelle in der Zelle an der angegebenen Zeile und Spalte zurück."""
        pass
    @abstractmethod
    def get_row_values(self, row: int) -> list:
        """Gibt die Werte aller Zellen in der angegebenen Zeile als Liste zurück."""
        pass
    @abstractmethod
    def get_cell_format(self, row: int, col: int) -> dict:
        """Gibt das Format der Zelle in der angegebenen Zeile und Spalte als Dictionary zurück."""
        pass
    @abstractmethod
    def get_row_formats(self, row: int) -> list:
        """Gibt die Formate aller Zellen in der angegebenen Zeile als Liste von Dictionaries zurück."""
        pass

    # Tester
    @abstractmethod
    def is_readonly(self) -> bool:
        """Gibt zurück, ob die Datei nur lesend ist."""
        pass

    def is_engine_readonly(self) -> bool:
        """Gibt zurück, ob die Engine nur lesend ist."""
        pass

    # Setter
    @abstractmethod
    def set_cell_value(self, row: int, col: int, value):
        """Setzt den Wert der Zelle in der angegebenen Zeile und Spalte."""
        pass
    @abstractmethod
    def add_row(self, row: int):
        """Setzt die Werte aller Zellen in der angegebenen Zeile."""
        pass
    @abstractmethod
    def set_row_values(self, row: int, values: list):
        """Setzt die Werte aller Zellen in der angegebenen Zeile."""
        pass

    @abstractmethod
    def set_cell_format(self, row: int, col: int, fmt: dict):
        """Setzt das Format der Zelle in der angegebenen Zeile und Spalte."""
        pass
    @abstractmethod
    def set_row_formats(self, row: int, formats: list):
        """Setzt die Formate aller Zellen in der angegebenen Zeile."""
        pass


class TableEnginePyxl(TableEngine):
    def __init__(self, ws):
        self.ws = ws

    def get_max_row(self) -> int:
        return self.ws.max_row

    def get_max_col(self) -> int:
        return self.ws.max_column

    def get_cell_value(self, row: int, col: int):
        return self.ws.cell(row=row, column=col).value

    def get_row_values(self, row: int) -> list:
        return [self.ws.cell(row=row, column=c).value for c in range(1, self.get_max_col() + 1)]

    def get_cell_format(self, row: int, col: int) -> dict:
        cell = self.ws.cell(row=row, column=col)
        font = cell.font
        fill = cell.fill
        return {
            "font_name": font.name,
            "font_size": font.size,
            "bold": font.bold,
            "italic": font.italic,
            "font_color": font.color.rgb if font.color and font.color.type == "rgb" else None,
            "fill_color": fill.fgColor.rgb if fill and fill.fgColor.type == "rgb" else None
        }

    def get_row_formats(self, row: int) -> list:
        return [self.get_cell_format(row, c) for c in range(1, self.get_max_col() + 1)]

    def is_readonly(self) -> bool:
        return self.ws.parent.read_only

    def is_engine_readonly(self) -> bool:
        return False

    # Setter
    def set_cell_value(self, row: int, col: int, value):
        self.ws.cell(row=row, column=col).value = value

    def add_row(self, row: int):
        self.ws.insert_rows(row)

    def set_row_values(self, row: int, values: list):
        for c, val in enumerate(values, start=1):
            self.set_cell_value(row, c, val)

    def set_cell_format(self, row: int, col: int, fmt: dict):
        from openpyxl.styles import Font, PatternFill
        cell = self.ws.cell(row=row, column=col)
        # Font
        cell.font = Font(
            name=fmt.get("font_name", cell.font.name),
            size=fmt.get("font_size", cell.font.size),
            bold=fmt.get("bold", cell.font.bold),
            italic=fmt.get("italic", cell.font.italic),
            color=fmt.get("font_color", cell.font.color)
        )
        # Fill
        if fmt.get("fill_color"):
            cell.fill = PatternFill(start_color=fmt["fill_color"], end_color=fmt["fill_color"], fill_type="solid")

    def set_row_formats(self, row: int, formats: list):
        for c, fmt in enumerate(formats, start=1):
            self.set_cell_format(row, c, fmt)



class TableEnginePyexcel(TableEngine):
    """
    TableEngine für .xls-Dateien auf Basis von pyexcel.

    pyexcel ist nur zum Lesen gedacht – Schreiben und Formatierung sind nicht unterstützt.
    Setter-Methoden werfen daher NotImplementedError.
    """

    def __init__(self, sheet):
        self.sheet = sheet

    def get_max_row(self) -> int:
        return self.sheet.number_of_rows()

    def get_max_col(self) -> int:
        return self.sheet.number_of_columns()

    def get_cell_value(self, row: int, col: int):
        try:
            # 0-based indexing im Sheet, wir arbeiten 1-based
            return self.sheet[row - 1, col - 1]
        except (IndexError, KeyError):
            return None

    def get_row_values(self, row: int) -> list:
        try:
            # sheet.row ist eine Liste von Row-Objekten, also per Index zugreifen
            raw_row = self.sheet.row[row - 1]
            # Row ist iterable, wir wandeln in eine einfache Liste um
            return list(raw_row)
        except (IndexError, KeyError):
            return []

    def get_cell_format(self, row: int, col: int) -> dict:
        # pyexcel liefert keine Formatinfos – wir geben "General" zurück
        return {"number_format": "General"}

    def get_row_formats(self, row: int) -> list:
        return [{"number_format": "General"} for _ in range(self.get_max_col())]

    def is_readonly(self) -> bool:
        return True

    def is_engine_readonly(self) -> bool:
        return True

    # Schreib-APIs nicht unterstützt
    def set_cell_value(self, row: int, col: int, value):
        raise NotImplementedError("pyexcel unterstützt kein Schreiben von Werten.")

    def add_row(self, row: int):
        raise NotImplementedError("pyexcel unterstützt kein Schreiben von Werten.")

    def set_row_values(self, row: int, values: list):
        raise NotImplementedError("pyexcel unterstützt kein Schreiben von Werten.")

    def set_cell_format(self, row: int, col: int, fmt: dict):
        raise NotImplementedError("pyexcel unterstützt kein Schreiben von Formaten.")

    def set_row_formats(self, row: int, formats: list):
        raise NotImplementedError("pyexcel unterstützt kein Schreiben von Formaten.")

# ============================================================
# Iteration
# ============================================================

class TableRowEnumerator:
    """
    Iterator für TableEngine-Zeilen.

    Ermöglicht zeilenweises Durchlaufen eines Worksheets via `next()` oder `for row in ...`.
    Zusätzlich kann mit `add_row(values)` eine neue Zeile an der aktuellen Position eingefügt werden.

    Gibt bei jedem Schritt ein Tupel `(row_index, row_values)` zurück.

    Beispiel 1:
        engine = load_engine("daten.xlsx", "Messwerte")
        for row_index, row_values in TableRowEnumerator(engine):
            print(f"Zeile {row_index}: {row_values}")

    Beispiel 1:
        enumerator = TableRowEnumerator(engine)
        while True:
            try:
                r, values = next(enumerator)
                # Verarbeitung...
                row_index = enum.add_row(["Messwert A", 42.0, True])
                print(f"Zeile {row_index} wurde eingefügt.")
            except StopIteration:
                break
    """

    def __init__(self, engine: TableEngine, start_row: int = 1):
        self.engine = engine
        self.current = start_row
        self.max_row = engine.get_max_row()

    def __iter__(self):
        return self

    def __next__(self):
        if self.current > self.max_row:
            raise StopIteration
        row_values = self.engine.get_row_values(self.current)
        result = (self.current, row_values)
        self.current += 1
        return result

    def add_row(self, values: list) -> int:
        """
        Fügt eine neue Zeile an der aktuellen Position ein.

        :param values: Liste der Zellwerte für die neue Zeile.
        :return: Der Index der eingefügten Zeile.
        """
        self.engine.add_row(self.current)
        self.engine.set_row_values(self.current, values)
        inserted_row = self.current
        self.current += 1
        self.max_row += 1
        return inserted_row

    def get_max_row(self):
        return self.max_row

# ============================================================
# Factory
# ============================================================

def load_engine(file_path: str, sheet_name: str) -> tuple[Any, TableEngine]:
    """
    Lädt eine Excel-Datei (.xlsx oder .xls) und gibt das Workbook sowie die zugehörige TableEngine zurück.

    Args:
        file_path (str): Pfad zur Excel-Datei.
        sheet_name (str): Name des zu ladenden Sheets.

    Returns:
        Tuple[Workbook, TableEngine]: Das Workbook-Objekt und die passende TableEngine-Instanz.

    Raises:
        ImportError: Falls xlrd für .xls nicht installiert ist.
        ValueError: Falls die Dateiendung nicht unterstützt wird.
    """

    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xlsx":
        wb = openpyxl.load_workbook(file_path, data_only=False)
        return wb, TableEnginePyxl(wb[sheet_name])
    elif ext == ".xls":
        try:
            import pyexcel
        except ImportError:
            raise ImportError("Das Paket 'pyexcel' ist nicht installiert. Installiere es mit 'pip install pyexcel'.")
        try:
            import pyexcel_xls
        except ImportError:
            raise ImportError("Das Paket 'pyexcel_xls' ist nicht installiert. Installiere es mit 'pip install pyexcel_xls'.")
        wb = pyexcel.get_book(file_name=file_path)
        return wb, TableEnginePyexcel(wb.sheet_by_name(sheet_name))
    elif ext == ".ods":
        try:
            import pyexcel
        except ImportError:
            raise ImportError("Das Paket 'pyexcel' ist nicht installiert. Installiere es mit 'pip install pyexcel'.")
        try:
            import pyexcel_ods
        except ImportError:
            raise ImportError("Das Paket 'pyexcel_ods' ist nicht installiert. Installiere es mit 'pip install pyexcel_ods'.")
        wb = pyexcel.get_book(file_name=file_path)
        return wb, TableEnginePyexcel(wb.sheet_by_name(sheet_name))
    else:
        raise ValueError(f"Nicht unterstützte Dateiendung: {ext}")
